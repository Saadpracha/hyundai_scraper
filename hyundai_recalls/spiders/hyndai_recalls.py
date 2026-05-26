import json
import os

import scrapy
from openpyxl import load_workbook
from scrapy.utils.project import get_project_settings
from twocaptcha import TwoCaptcha


def _clean(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


_STATUS_LABELS = {
    "en": {
        "text-complete": "Complete",
        "text-incomplete": "Incomplete",
    },
    "fr": {
        "text-complete": "Complété",
        "text-incomplete": "Incomplet",
    },
}

_I18N_LABELS = {
    "en": {
        **_STATUS_LABELS["en"],
        "text-safety-recall": "Safety Recall",
        "text-emission-recall": "Emission Recall",
    },
    "fr": {
        **_STATUS_LABELS["fr"],
        "text-safety-recall": "Rappels de sécurité",
        "text-emission-recall": "Rappel d'émission",
    },
}


def _text_or_i18n(element, labels):
    if element is None:
        return None
    text = _clean(element.xpath("string(.)").get())
    if text:
        return text
    key = element.xpath("@data-i18n").get()
    return labels.get(key) if key else None


def _find_column_index(header_lower, *keywords):
    for keyword in keywords:
        for i, name in enumerate(header_lower):
            if keyword in name:
                return i
    return None


def _load_input_from_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return [], None

    header = [_clean(cell) or "" for cell in rows[0]]
    header_lower = [h.lower() for h in header]

    if any("vin" in name for name in header_lower):
        vin_idx = _find_column_index(header_lower, "vin")
        data_rows = rows[1:]
    else:
        vin_idx = 0
        data_rows = rows

    account_idx = _find_column_index(
        header_lower, "account", "dealership", "dealer"
    )

    vins = []
    account = None
    seen = set()
    for row in data_rows:
        if not row or vin_idx >= len(row):
            continue

        if account is None and account_idx is not None and account_idx < len(row):
            account = _clean(row[account_idx])

        vin = _clean(row[vin_idx])
        if not vin:
            continue
        vin = vin.upper()
        if vin not in seen:
            seen.add(vin)
            vins.append(vin)
    return vins, account


def _parse_vehicle(response):
    raw = response.xpath(
        '//h3[contains(text(),"Vehicle Information")]/following-sibling::p[1]/text()'
    ).get()
    if not raw:
        raw = response.xpath(
            '//h3[@data-i18n="heading-vehicle-information"]/following-sibling::p[1]/text()'
        ).get()

    raw = _clean(raw)
    if not raw:
        return None, None

    parts = raw.split(None, 1)
    year = parts[0]
    model = _clean(parts[1]) if len(parts) > 1 else None
    return year, model


def _parse_warranty_start_date(response):
    return _clean(
        response.xpath(
            '//label[@data-i18n="warranty-start-date"]/following-sibling::text()'
        ).get()
    )


def _parse_recall_panel(panel, labels):
    summary = panel.xpath('./div[@class="panel-body"][1]')

    title_parent = panel.xpath(
        './/span[@data-i18n="text-safety-recall"]/parent::div'
    )
    if not title_parent:
        title_parent = panel.xpath(
            './/span[@data-i18n="text-emission-recall"]/parent::div'
        )

    description_parts = panel.xpath(
        './/h4[@data-i18n="text-description"]/following-sibling::span//text()'
    ).getall()

    safety_recall_el = panel.xpath('.//span[@data-i18n="text-safety-recall"]')
    emission_recall_el = panel.xpath('.//span[@data-i18n="text-emission-recall"]')

    if emission_recall_el:
        recall_type = "emission_recall"
    elif safety_recall_el:
        recall_type = "safety_recall"
    else:
        recall_type = None

    return {
        "recall_type": recall_type,
        "status": _text_or_i18n(
            summary.xpath('.//strong[@data-i18n]')[0]
            if summary.xpath('.//strong[@data-i18n]')
            else None,
            labels,
        ),
        "date_repaired": _clean(
            summary.xpath(
                './/div[contains(@class,"col-sm-2") and contains(@class,"hidden-xs")]/text()'
            ).get()
        ),
        "recall_number": _clean(
            summary.xpath('.//div[contains(@class,"col-sm-3")]/text()').get()
        ),
        "transport_canada_number": _clean(
            summary.xpath('.//div[contains(@class,"col-sm-4")]/text()').get()
        ),
        "title": _clean(
            title_parent.xpath('./text()[normalize-space()][1]').get()
        ),
        "description": _clean(
            " ".join(part.strip() for part in description_parts if part.strip())
        ),
    }


def _build_standard_item(
    vin,
    year,
    model,
    source_url,
    language,
    recall,
    warranty_start_date=None,
    make="hyundai",
    account=None,
):
    other_fields = {}
    if warranty_start_date:
        other_fields["warranty_start_date"] = warranty_start_date
    if recall.get("date_repaired"):
        other_fields["date_repaired"] = recall["date_repaired"]
    if recall.get("transport_canada_number"):
        other_fields["transport_canada_number"] = recall["transport_canada_number"]

    return {
        "account": account,
        "vin": vin,
        "year": year,
        "make": make,
        "model": model,
        "type": recall.get("recall_type"),
        "campaign": recall.get("recall_number"),
        "status": recall.get("status"),
        "source_url": source_url,
        "language": language,
        "title": recall.get("title"),
        "description": recall.get("description"),
        "other_fields": json.dumps(other_fields) if other_fields else None,
    }


class HyndaiRecallsSpider(scrapy.Spider):
    name = "hyndai_recalls"
    allowed_domains = ["recall.hyundaicanada.com"]
    make = "hyundai"

    custom_settings = {
        "ROBOTSTXT_OBEY": False,
    }

    def __init__(
        self,
        vin_number=None,
        captcha_api_key=None,
        lang=None,
        input_file=None,
        account=None,
        *args,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)

        self.account = _clean(account)
        self.lang = (lang or "en").lower()
        if self.lang not in _I18N_LABELS:
            raise ValueError("lang must be 'en' or 'fr'")

        self.i18n_labels = _I18N_LABELS[self.lang]
        self.base_url = (
            "https://recall.hyundaicanada.com/"
            if self.lang == "en"
            else "https://recall.hyundaicanada.com/fr"
        )
        self.results_url = f"{self.base_url.rstrip('/')}/Home/Results"

        project_root = os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        )
        input_path = input_file or os.path.join(
            project_root, "hyndai_recalls_input.xlsx"
        )

        if vin_number:
            self.vin_numbers = [_clean(vin_number).upper()]
        elif os.path.isfile(input_path):
            self.vin_numbers, file_account = _load_input_from_xlsx(input_path)
            if not self.account:
                self.account = file_account
            self.logger.info(
                "Loaded %d VIN(s) from %s", len(self.vin_numbers), input_path
            )
            if self.account:
                self.logger.info("Dealership account: %s", self.account)
        else:
            raise ValueError(
                f"No VINs found. Pass -a vin_number=... or provide {input_path}"
            )

        if not self.vin_numbers:
            raise ValueError("No VIN numbers to process")

        settings = get_project_settings()
        self.captcha_api_key = (
            captcha_api_key
            or os.environ.get("TWOCAPTCHA_API_KEY")
            or settings.get("TWOCAPTCHA_API_KEY")
        )
        self._vin_index = 0

    def start_requests(self):
        yield scrapy.Request(
            self.base_url,
            callback=self.parse,
            meta={"vin": self.vin_numbers[0]},
            dont_filter=True,
        )

    def _next_vin_request(self):
        self._vin_index += 1
        if self._vin_index >= len(self.vin_numbers):
            return None

        vin = self.vin_numbers[self._vin_index]
        self.logger.info(
            "Processing VIN %d/%d: %s",
            self._vin_index + 1,
            len(self.vin_numbers),
            vin,
        )
        return scrapy.Request(
            self.base_url,
            callback=self.parse,
            meta={"vin": vin},
            dont_filter=True,
        )

    def parse(self, response):
        vin = response.meta["vin"]
        self.logger.info(
            "Processing VIN %d/%d: %s",
            self._vin_index + 1,
            len(self.vin_numbers),
            vin,
        )

        site_key = response.xpath(
            '//div[@class="g-recaptcha"]/@data-sitekey'
        ).get()
        verification_token = response.xpath(
            '//input[@name="__RequestVerificationToken"]/@value'
        ).get()

        if not site_key:
            self.logger.error("Captcha sitekey not found for VIN %s", vin)
            next_request = self._next_vin_request()
            if next_request:
                yield next_request
            return

        if not verification_token:
            self.logger.error("Verification token not found for VIN %s", vin)
            next_request = self._next_vin_request()
            if next_request:
                yield next_request
            return

        if not self.captcha_api_key:
            self.logger.error(
                "Set TWOCAPTCHA_API_KEY in settings, environment, or -a captcha_api_key"
            )
            return

        solver = TwoCaptcha(self.captcha_api_key)

        try:
            result = solver.recaptcha(sitekey=site_key, url=response.url)
            captcha_token = result.get("code")
            self.logger.info("CAPTCHA solved for VIN %s", vin)
        except Exception as e:
            self.logger.error("Captcha solve failed for VIN %s: %s", vin, e)
            next_request = self._next_vin_request()
            if next_request:
                yield next_request
            return

        yield scrapy.FormRequest(
            url=self.results_url,
            formdata={
                "__RequestVerificationToken": verification_token,
                "VINNumber": vin,
                "g-recaptcha-response": captcha_token,
            },
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Origin": "https://recall.hyundaicanada.com",
                "Referer": self.base_url,
            },
            callback=self.parse_results,
            meta={"vin": vin},
            dont_filter=True,
        )

    def parse_results(self, response):
        vin = response.meta["vin"]
        self.logger.info("Results for VIN %s: HTTP %s", vin, response.status)

        year, model = _parse_vehicle(response)
        warranty_start_date = _parse_warranty_start_date(response)

        panels = response.xpath('//div[@class="panel panel-default"]')
        if not panels:
            yield _build_standard_item(
                vin,
                year,
                model,
                response.url,
                self.lang,
                {},
                warranty_start_date,
                self.make,
                self.account,
            )

        for panel in panels:
            recall = _parse_recall_panel(panel, self.i18n_labels)
            yield _build_standard_item(
                vin,
                year,
                model,
                response.url,
                self.lang,
                recall,
                warranty_start_date,
                self.make,
                self.account,
            )

        next_request = self._next_vin_request()
        if next_request:
            yield next_request
